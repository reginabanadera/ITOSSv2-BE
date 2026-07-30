from openpyxl import load_workbook
import uuid
import os
from datetime import datetime
from bs4 import BeautifulSoup

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

output_dir = os.path.join(BASE_DIR, "..", "generated")
output_dir = os.path.normpath(output_dir)

os.makedirs(output_dir, exist_ok=True)

def fill_excel_template(template_name, template_path, fields):

    wb = load_workbook(template_path)
    ws = wb.active

    if template_name == "UFSCreation":
        fill_ufs_creation(ws, fields)

    elif template_name == "EmailCreation":
        fill_email_creation(ws, fields)
    
    elif template_name == "GroupEmailCreation":
        fill_grpemail_creation(ws, fields)

    elif template_name == "URSCreation":
        fill_urs_creation(ws, fields)

    elif template_name == "DataPatch":
        fill_dpatch_creation(ws, fields)

    filename = f"{uuid.uuid4()}.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb.save(output_path)

    return output_path

def fill_ufs_creation(ws, fields):
    first_name = fields.get("FirstName", "")
    last_name = fields.get("LastName", "")
    emailaddress = fields.get("EmailAddress", "")
    terminals = fields.get("Terminal", [])

    start_row = 12

    for index, terminal in enumerate(terminals):
        row = start_row + index

        ws[f"B{row}"] = terminal.get("RequestType", "")
        ws[f"E{row}"] = first_name
        ws[f"F{row}"] = last_name
        ws[f"G{row}"] = terminal.get("TerminalId", "")
        ws[f"J{row}"] = emailaddress
        ws[f"L{row}"] = terminal.get("SecurityLevel", "")
        ws[f"P{row}"] = terminal.get("HighSecurityUser", "")


def fill_email_creation(ws, fields):
    first_name = fields.get("FirstName", "").title()
    last_name = fields.get("LastName", "").title()
    full_name = f"{first_name} {last_name}"
    email_address = fields.get("EmailAddress", "")
    date_approved = fields.get("DateActed", "")

    dt = datetime.fromisoformat(date_approved)

    formatted = dt.strftime("%b. %d, %Y")

    ws[f"C5"] = "Rodger Corella"
    ws[f"C6"] = formatted   
    ws[f"D11"] = first_name
    ws[f"D13"] = last_name
    ws[f"D14"] = full_name
    ws[f"D15"] = email_address

    ws[f"D20"] = "SEO"
    ws[f"D21"] = "Philippines"
    ws[f"D22"] = "KWEPH"

def fill_grpemail_creation(ws, fields):
    requestType = fields.get("RequestType", "")
    groupName = fields.get("GroupName", "")
    groupEmail = fields.get("GroupEmailAddress", "")
    
    members = fields.get("Member", [])
    member_text = "\n".join(
        member["Member"] for member in members
    )

    owners = fields.get("Owner", [])
    owner_text = "\n".join(
        owner["Owner"] for owner in owners
    )

    ws[f"D10"] = requestType
    ws[f"D12"] = groupName   
    ws[f"D14"] = groupEmail
    ws[f"D16"] = "SEO"
    ws[f"D17"] = "Philippines"
    ws[f"D18"] = "KWEPH"

    ws[f"D20"] = owner_text
    ws[f"D21"] = member_text


def fill_urs_creation(ws, fields):
    first_name = fields.get("FirstName", "")
    last_name = fields.get("LastName", "")
    terminals = fields.get("Terminal", [])

    start_row = 14

    for index, terminal in enumerate(terminals):
        row = start_row + index

        ws[f"B{row}"] = terminal.get("RequestType", "")
        ws[f"E{row}"] = first_name
        ws[f"F{row}"] = last_name
        ws[f"G{row}"] = terminal.get("TerminalId", "")
        ws[f"H{row}"] = terminal.get("Authority", "")
        ws[f"I{row}"] = terminal.get("LoginModeUFSAE", "")
        ws[f"J{row}"] = terminal.get("LoginModeUFSAI", "")
        ws[f"K{row}"] = terminal.get("LoginModeUFSOE", "")
        ws[f"L{row}"] = terminal.get("LoginModeUFSOI", "")
        ws[f"M{row}"] = terminal.get("LoginModeUAS", "")
        ws[f"N{row}"] = terminal.get("LoginModeProfile", "")
        ws[f"O{row}"] = terminal.get("SecurityLevel", "")
        ws[f"P{row}"] = "570"

def fill_dpatch_creation(ws, fields):
    system_name = fields.get("SystemName", "")
    current_user = fields.get("CurrentUser", "").title()
    date_today = fields.get("DateToday", "")

    ws[f"E12"] = date_today
    ws[f"E15"] = system_name
    ws[f"E18"] = current_user

    html = fields.get("Description", "")

    soup = BeautifulSoup(html, "html.parser")

    description = soup.get_text("\n", strip=True)

    ws["C27"] = description

